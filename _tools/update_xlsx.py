# -*- coding: utf-8 -*-
"""
Update マイルストーン1_要件整理.xlsx to reflect:
- Actual MS2 implementation (MIKOMERU-style 3-group sidebar, settings page added, LP minimum)
- Sidebar-structure impact on MS3-MS6 screen naming
Scope is kept as-is (no feature expansion beyond already-approved sender template).
"""
import sys, io, copy, shutil, pathlib
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy as style_copy

SRC = pathlib.Path("マイルストーン1_要件整理.bak.xlsx")

wb = openpyxl.load_workbook(SRC, data_only=False)

# ---------- helpers ----------
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="E6F0FA")
SECTION_FILL = PatternFill("solid", fgColor="F2F2F2")
BOLD = Font(bold=True)

def append_section(ws, title, rows, row_cursor, bold_header_row=True, header_fill=True):
    """Append a merged section title + data rows starting from row_cursor.
    `rows` is a list of lists (1 row per entry); the first row is treated as header if bold_header_row True.
    """
    # Section title (merged across used columns, based on widest row)
    width = max(1, max((len(r) for r in rows), default=1))
    ws.cell(row=row_cursor, column=1, value=title).font = Font(bold=True, size=11, color="1F2A34")
    ws.cell(row=row_cursor, column=1).fill = SECTION_FILL
    if width > 1:
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=width)
    row_cursor += 1

    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            cell = ws.cell(row=row_cursor, column=j + 1, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
            if i == 0 and bold_header_row:
                cell.font = BOLD
                if header_fill:
                    cell.fill = HEADER_FILL
        row_cursor += 1

    return row_cursor + 1  # blank row after section


def next_empty_row(ws):
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            return r + 2
    return 1


# ================================================================
# 03_最小MVP機能一覧
# ================================================================
ws = wb["03_最小MVP機能一覧"]
# Fix F-04 terminology: 送信中 -> 配信中
if ws["E7"].value and "送信中" in str(ws["E7"].value):
    ws["E7"].value = "準備中 → 配信中 → 完了 / 停止"
    print("[03] F-04 E7 updated: 送信中 -> 配信中")

# Extend F-09 (テンプレート CRUD) scope to include sender template (独立テンプレとして追加済み)
# Find F-09 row
f09_row = None
for r in ws.iter_rows(min_col=1, max_col=1):
    if r[0].value == "F-09":
        f09_row = r[0].row
        break
if f09_row:
    ws.cell(row=f09_row, column=5).value = (
        "送信文章テンプレ: 件名・本文・変数 {{会社名}}, 送信元テンプレ (差出人プロファイル): 会社名・氏名・メール・電話を再利用"
    )
    print(f"[03] F-09 E{f09_row} updated: 送信元テンプレ含めるよう変更")

# Append new features (sidebar items not originally listed)
cursor = next_empty_row(ws)
cursor = append_section(
    ws,
    "■ 追加機能（サイドバー項目対応・要件xlsx 未反映分を補足）",
    [
        ["No", "カテゴリ", "機能", "対応MS", "最小実装の中身", "MIKOMERU 比較メモ"],
        [
            "F-21",
            "テンプレート",
            "送信元テンプレート CRUD（差出人プロファイル）",
            "MS3",
            "会社名・担当者氏名・メール・電話・郵便番号・住所・URL を登録し、送信時に選択して入力欄を自動フィル。送信文章テンプレからも参照可能。",
            "MIKOMERU の『送信元テンプレート』相当。白石様との合意で本 MVP に含む。",
        ],
        [
            "F-22",
            "設定",
            "アカウント / システム設定",
            "MS2 / MS3",
            "MS2: アカウント情報表示（表示名・メール）。MS3: 文字数プリセット（200字/500字等）・CSV 既定項目・送信元テンプレ既定値の編集。",
            "MIKOMERU 元は別メニュー。ログイン後の確認導線として最小実装。",
        ],
        [
            "F-23",
            "ホーム",
            "ホーム（統計情報サマリ画面）",
            "MS2 / MS5",
            "MS2: 進行中案件数・今月新規案件数・KPI プレースホルダ。MS5: 送信数・成功率 KPI・日次/案件別グラフ・最新ジョブログを注入。",
            "MIKOMERU の『ホーム』相当。MS2 では骨格、MS5 で中身投入。",
        ],
    ],
    cursor,
)

# Append MS2-reflection section
cursor = next_empty_row(ws)
cursor = append_section(
    ws,
    "■ MS2 実装時点の反映（サイドバー構成・UI 最小実装）",
    [
        ["No", "項目", "内容", "対応MS"],
        [
            "U-01",
            "サイドバー構成",
            "MIKOMERU 準拠の 3 グループ構造（フォーム送信 / 会社情報 / その他）＋ ホーム・案件はトップレベル。MS3 以降で有効化する項目は『予定』バッジで非活性表示。",
            "MS2",
        ],
        [
            "U-02",
            "共通シェル",
            "左サイドバー + ヘッダ（ログイン中ユーザ表示・ログアウト）+ パンくず。全画面で共通。",
            "MS2",
        ],
        [
            "U-03",
            "ホーム (P02) 骨格",
            "進行中案件件数 / 今月新規案件件数 を表示。KPI カード（送信数・成功率）は MS5 で中身投入するプレースホルダ。",
            "MS2",
        ],
        [
            "U-04",
            "設定 (P13) 最小",
            "アカウント情報表示。文字数プリセット・CSV 既定は MS3 で編集可能化。",
            "MS2",
        ],
        [
            "U-05",
            "LP (:3001) 最小",
            "Next.js で 1 ページ構成確認用。本実装 (P14) は MS6。",
            "MS2",
        ],
    ],
    cursor,
)

# ================================================================
# 04_画面一覧
# ================================================================
ws = wb["04_画面一覧"]
# Soften the "除外" note (P13/LP are no longer fully excluded)
for row in ws.iter_rows(min_col=1, max_col=1):
    c = row[0]
    if c.value and "設定画面・LP は最小 MVP からは除外" in str(c.value):
        c.value = (
            "※ 設定画面 (P13) は MS2 で最小実装済み。LP (P14) は MS2 で :3001 最小実装、"
            "本実装は MS6。サイドバー構成は下部『サイドバー構成 (MIKOMERU 準拠 3 グループ)』参照。"
        )
        print(f"[04] {c.coordinate} 除外ノート更新")
        break

# Append P13 and P14 rows + sidebar structure
cursor = next_empty_row(ws)
cursor = append_section(
    ws,
    "■ 追加画面 (MS2 実装時点)",
    [
        ["ID", "画面名", "対応MS", "主要要素", "備考"],
        ["P13", "設定", "MS2", "アカウント情報 / MS3プレースホルダ", "最小実装"],
        ["P14", "LP", "MS2", "サービス紹介 1 ページ", ":3001, 本実装は MS6"],
    ],
    cursor,
)

cursor = append_section(
    ws,
    "■ サイドバー構成 (MIKOMERU 準拠 3 グループ) + アイコン + MS 別実装内容",
    [
        ["グループ", "サイドバーラベル", "アイコン", "リンク先", "有効化MS", "MS 別実装内容", "備考"],
        ["—", "ホーム", "家 (home)", "P02 (骨格)",
         "MS2 / MS5",
         "MS2: 進行中案件・今月新規案件・サマリの骨格 / MS5: 送信 KPI カード・日次/案件別グラフ・最新ジョブログ",
         "KPI は MS5 で本実装"],
        ["—", "案件", "ブリーフケース (case)", "P03 → P04 / P05",
         "MS2",
         "案件 CRUD（登録・一覧・詳細・編集・複製・論理/物理削除）・検索・絞込・ステータス遷移（準備中/配信中/完了/停止）",
         "—"],
        ["フォーム送信", "自動送信", "紙飛行機 (send)", "P11 送信キュー",
         "MS4",
         "送信ジョブ作成・承認・実行（Playwright 汎用フォーム送信）、キュー管理（pending/running/done/failed）、一時停止/再開/キャンセル、BL 自動除外、レート制限",
         "MS2-3 は非活性（予定バッジ）"],
        ["フォーム送信", "自動送信ログ", "ログ一覧 (log)", "P12 配信結果一覧",
         "MS5",
         "送信結果一覧（案件・期間・成否・エラー種別で絞込）、結果詳細モーダル（HTMLスナップショット閲覧）、CSV エクスポート、エラー種別 5 分類",
         "MS2-4 は非活性"],
        ["フォーム送信", "送信文章テンプレート", "ドキュメント (docText)", "P09 / P10",
         "MS3",
         "テンプレート CRUD（検索・複製・削除）、件名・本文編集、変数 {{会社名}} {{担当者名}} 対応、プレビュー・文字数カウンタ、案件との紐付け",
         "MS2 は非活性"],
        ["フォーム送信", "送信元テンプレート", "ユーザー (user)", "(新) 差出人プロファイル CRUD",
         "MS3",
         "送信元プロファイル CRUD（会社名・担当者氏名・メール・電話・郵便番号・住所・URL）、送信時に選択して入力欄自動フィル、送信文章テンプレから参照",
         "MS2 は非活性、F-09 拡張 / MIKOMERU 準拠の『送信元テンプレート』相当"],
        ["フォーム送信", "送信除外設定", "ブロック (block)", "P08 ブラックリスト",
         "MS3",
         "ブラックリスト CRUD（ドメイン or 会社名 / 理由 / 登録日）、手動追加 + 取込時自動追加（『営業お断り』等キーワード検出）、送信直前の BL 突合",
         "MS2 は非活性"],
        ["会社情報", "リスト取込", "アップロード (upload)", "P06",
         "MS3",
         "CSV アップロード UI（UTF-8 / Shift_JIS 自動判定）、項目マッピング UI（会社名・form_url 必須）、取込実行（重複排除・BL 突合・エラー行レポート）",
         "MS2 は非活性"],
        ["会社情報", "保存済みリスト", "リスト (list)", "P07",
         "MS3",
         "リスト一覧（各リストの企業数表示）、リスト詳細（企業行の編集・削除）、案件との紐付け",
         "MS2 は非活性"],
        ["その他", "設定", "歯車 (cog)", "P13",
         "MS2 / MS3",
         "MS2: アカウント情報表示（表示名・メール） / MS3: 文字数プリセット（200字/500字等）・CSV 既定項目設定・送信元テンプレートの既定値",
         "MS2 最小実装済み"],
    ],
    cursor,
)

# ================================================================
# MS2_案件管理
# ================================================================
ws = wb["MS2_案件管理"]

# Update tasks 9, 10
for r in range(10, 22):
    v = ws.cell(row=r, column=2).value
    if v == "案件削除（論理削除）":
        ws.cell(row=r, column=2).value = "案件削除 (物理削除) + 複製"
        ws.cell(row=r, column=3).value = "deleteCase / duplicateCase Server Action"
        ws.cell(row=r, column=4).value = "0.25 日"
        print(f"[MS2] task 9 row{r} updated")
    if v == "Cypress/Playwright によるスモークテスト":
        ws.cell(row=r, column=2).value = "疎通確認 (curl + tsc typecheck)"
        ws.cell(row=r, column=3).value = "ログイン→全画面 HTTP 200 確認、型エラー 0 件"
        ws.cell(row=r, column=4).value = "0.25 日"
        print(f"[MS2] task 10 row{r} updated (E2E 自動化は MS6 へ)")

cursor = next_empty_row(ws)

# Additional tasks section (reflect actual MS2 work)
cursor = append_section(
    ws,
    "■ 追加実装タスク (MS2 完了時点で反映)",
    [
        ["No", "タスク", "成果物・出力", "見積目安"],
        ["A-1", "モノレポ構造 (npm workspaces) + Next.js 15 App Router + Prisma", "apps/admin, apps/lp, packages/db", "0.5 日"],
        ["A-2", "共通シェル (MIKOMERU 式 3 グループサイドバー + ヘッダ + パンくず)", "Sidebar/Header/Breadcrumbs コンポーネント", "0.5 日"],
        ["A-3", "ホーム画面 (P02) 骨格 – KPI カード 3 枚プレースホルダ + サマリ", "/home ページ (進行中案件・今月新規件数は実動)", "0.5 日"],
        ["A-4", "設定画面 (P13) 最小 – アカウント情報表示", "/settings ページ", "0.25 日"],
        ["A-5", "LP (:3001) 最小 1 ページ", "apps/lp (構成確認用)", "0.25 日"],
        ["A-6", "GitHub リポジトリ作成・初回プッシュ", "origin/main", "0.1 日"],
    ],
    cursor,
)

# Additional deliverables
cursor = append_section(
    ws,
    "■ 追加納品物 (MS2 完了時点で反映)",
    [
        ["納品物", "備考"],
        ["ログイン画面 P01", "bcrypt + jose JWT セッション"],
        ["共通シェル", "MIKOMERU 式 3 グループサイドバー + ヘッダ + パンくず"],
        ["ホーム P02 骨格", "KPI プレースホルダは MS5 で中身投入"],
        ["設定 P13 最小", "アカウント情報 / 未実装項目はプレースホルダ"],
        ["LP :3001 最小", "本実装は MS6"],
        ["リポジトリ初期化", "monorepo + Prisma + DB migrate + seed"],
    ],
    cursor,
)

# Additional confirmation items
cursor = append_section(
    ws,
    "■ 追加確認事項 (MS2 完了時点で反映)",
    [
        ["確認項目", "判定方法"],
        ["MIKOMERU 式 3 グループサイドバーで全体構成が視認できる", "サイドバーで MS3 以降の予定項目が『予定』バッジ付きで確認できる"],
        ["認証後のみ管理画面にアクセス可能", "未ログインで /cases アクセス → /login リダイレクト"],
        ["案件の複製が動作する", "既存案件 → 複製 → 『(コピー)』が作成される"],
    ],
    cursor,
)

# ================================================================
# MS3_リスト_テンプレ
# ================================================================
ws = wb["MS3_リスト_テンプレ"]
cursor = next_empty_row(ws)
cursor = append_section(
    ws,
    "■ MS2 実装反映 (サイドバー構成・送信元テンプレ独立)",
    [
        ["No", "追加タスク", "成果物・出力", "見積目安"],
        ["B-1", "サイドバー 5 項目を有効化 (リスト取込 / 保存済みリスト / 送信文章テンプレ / 送信元テンプレ / 送信除外設定)", "Sidebar.tsx の disabled フラグを解除", "0.1 日"],
        ["B-2", "送信元テンプレート CRUD (差出人プロファイル: 会社名・氏名・メール・電話・郵便番号・住所)", "/templates/sender ページ + senderTemplates テーブル", "1 日"],
        ["B-3", "送信文章テンプレで送信元テンプレを選択 → プレビューに反映", "テンプレエディタ拡張", "0.25 日"],
    ],
    cursor,
)
cursor = append_section(
    ws,
    "■ 追加納品物",
    [
        ["納品物", "備考"],
        ["送信元テンプレート管理画面", "MIKOMERU の『送信元テンプレート』相当。実運用で入力負荷を大幅軽減"],
    ],
    cursor,
)

# ================================================================
# MS4_送信機能
# ================================================================
ws = wb["MS4_送信機能"]
cursor = next_empty_row(ws)
cursor = append_section(
    ws,
    "■ MS2 実装反映 (サイドバー構成)",
    [
        ["No", "追加タスク", "成果物・出力", "見積目安"],
        ["C-1", "サイドバー『自動送信』項目を有効化 (= P11 送信キュー)", "Sidebar.tsx", "0.05 日"],
        ["C-2", "P11 に『リストで送信 / CSVで送信』タブ切替 (MIKOMERU 準拠)", "タブコンポーネント", "0.25 日"],
        ["C-3", "送信元テンプレート選択で入力欄自動フィル (MS3 と連携)", "送信元テンプレ連携", "0.25 日"],
    ],
    cursor,
)

# ================================================================
# MS5_結果_ダッシュ
# ================================================================
ws = wb["MS5_結果_ダッシュ"]
# Update MS2 reference: P02 骨格が存在済み
for r in range(10, 22):
    v = ws.cell(row=r, column=2).value
    if v and "ダッシュボード画面（P02）" in str(v):
        ws.cell(row=r, column=2).value = (
            "ダッシュボード画面 (P02)：MS2 で作成した骨格を拡張 – 今月送信 / 成功 / 失敗 / 成功率 KPI"
        )
        print(f"[MS5] row{r} P02 骨格拡張に更新")

cursor = next_empty_row(ws)
cursor = append_section(
    ws,
    "■ MS2 実装反映 (サイドバー構成・ホーム骨格)",
    [
        ["No", "追加タスク", "成果物・出力", "見積目安"],
        ["D-1", "サイドバー『自動送信ログ』項目を有効化 (= P12 配信結果一覧)", "Sidebar.tsx", "0.05 日"],
        ["D-2", "P02 の進行中案件・今月新規案件サマリを残しつつ KPI を追加", "ホーム拡張", "0.25 日"],
        ["D-3", "エラー種別 5 分類 (成功 / 失敗 / 営業拒否 / フォームなし / キャンセル) を採用", "バッジ UI + 集計", "0.25 日"],
    ],
    cursor,
)

# ================================================================
# MS6_結合_納品
# ================================================================
ws = wb["MS6_結合_納品"]
cursor = next_empty_row(ws)
cursor = append_section(
    ws,
    "■ MS2 実装反映",
    [
        ["No", "追加タスク", "成果物・出力", "見積目安"],
        ["E-1", "LP (:3001) 構成確認ページ拡張 (MS2 の最小実装 → 本実装)", "apps/lp", "0.5 日"],
        ["E-2", "サイドバー全項目の有効化 + 予定バッジ除去", "Sidebar.tsx", "0.1 日"],
    ],
    cursor,
)

# ================================================================
# save
# ================================================================
wb.save(SRC)
print(f"\n[saved] {SRC}")
