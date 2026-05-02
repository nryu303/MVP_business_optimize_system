# -*- coding: utf-8 -*-
"""
AlphaScope案件 提案資料生成スクリプト
佐藤大将様向け 投資情報分析AI MVP提案
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = r"c:\MVP\milestone1\AlphaScope_提案資料.xlsx"

# ===== スタイル定義 =====
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Yu Gothic", size=12, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SUBHEADER_FONT = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
SECTION_FONT = Font(name="Yu Gothic", size=11, bold=True, color="000000")
NORMAL_FONT = Font(name="Yu Gothic", size=10)
BOLD_FONT = Font(name="Yu Gothic", size=10, bold=True)
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
RIGHT = Alignment(wrap_text=True, vertical="center", horizontal="right")


def apply_table(ws, start_row, headers, rows, col_widths=None, freeze=True):
    """ヘッダ＋データ行のテーブルを書き出し、罫線・色・幅を整える"""
    # ヘッダ
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    # データ
    for r, row in enumerate(rows, start_row + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP
            cell.border = BORDER
    # 列幅
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def write_title(ws, title, subtitle=None, span=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = HEADER_FILL
    c.font = Font(name="Yu Gothic", size=14, bold=True, color="FFFFFF")
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        s = ws.cell(row=2, column=1, value=subtitle)
        s.fill = SECTION_FILL
        s.font = SECTION_FONT
        s.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[2].height = 20


wb = Workbook()

# ============================================================
# Sheet 1: 表紙・サマリー
# ============================================================
ws = wb.active
ws.title = "1.サマリー"
write_title(
    ws,
    "AlphaScope - 投資情報分析AI MVP 提案資料",
    "佐藤大将様向け / 提案者: daiske / 2026-04-30",
)

summary = [
    ["項目", "内容"],
    ["プロジェクト名", "AlphaScope (仮称) - AI投資情報分析SaaS"],
    ["フェーズ", "MVP構築 → 初期ユーザー検証 (10名)"],
    ["開発期間", "約2週間 (集中開発) ＋ 1週間 (改善バッファ)"],
    ["開発費用 (人件費)", "20万円 (基本) / 30万円 (改善フェーズ込み)"],
    ["月額運用コスト概算 (MVP段階)", "約 1.5万円 〜 4万円 / 月"],
    ["月額運用コスト概算 (β版・100名規模)", "約 5万円 〜 15万円 / 月"],
    ["主要技術", "Python (FastAPI) / OpenAI API / Discord Bot / PostgreSQL / VPS or AWS"],
    ["最終成果物", "ニュース自動収集 → AI要約・スコアリング → Discord配信 の自動パイプライン"],
    ["将来展望", "配信先抽象化 (LINE/X) / 補助金申請用エビデンス / β版・全銘柄AI化"],
]
for r, row in enumerate(summary, 4):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = WRAP
        if r == 4:
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
            cell.alignment = CENTER
        else:
            cell.font = BOLD_FONT if c == 1 else NORMAL_FONT
            if c == 1:
                cell.fill = SECTION_FILL
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 80

# クライアント質問への回答セクション
ws.cell(row=16, column=1, value="■ クライアントからの質問への回答 (会話で求められた3点)").font = Font(
    name="Yu Gothic", size=12, bold=True, color="1F4E78"
)
qa = [
    ["No.", "ご質問", "回答 (要旨)"],
    [
        "1",
        "総額費用の概算 (運用コスト：サーバー・API等)",
        "MVP段階で月1.5〜4万円。β版100名規模で月5〜15万円。本番フル稼働で月20〜50万円を想定。"
        "内訳はサーバー(VPS or AWS) 月5千〜2万円、OpenAI API 月1万〜3万円、ニュースAPI 月0〜10万円(無料tier+スクレイピング併用なら抑制可)、"
        "DB/ストレージ 月3千〜1万円。詳細は『5.運用コスト見積』シート参照。",
    ],
    [
        "2",
        "人件費としての開発費用",
        "MVP一式で20万円 (募集記事ベース・1〜2週間集中開発)。要件整理・仕様調整・改善フェーズまで含めた場合は+10万円で計30万円が現実的。"
        "PM・要件整理・仕様書作成・実装・テスト・初期ユーザー対応までを1名で完遂する想定。",
    ],
    [
        "3",
        "実装までのスケジュール感",
        "Week1: 要件整理・データ収集設計・スクレイピング/API実装。"
        "Week2: AI処理(要約・スコアリング)・Discord Bot配信・10名テスト準備。"
        "Week3 (任意): バグ修正・配信頻度調整・改善。"
        "詳細は『7.マイルストーン』シート参照。",
    ],
]
for r, row in enumerate(qa, 18):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = WRAP
        if r == 18:
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
            cell.alignment = CENTER
        else:
            cell.font = NORMAL_FONT
ws.merge_cells(start_row=16, start_column=1, end_row=16, end_column=3)
# 質問テーブルだけ列構成が違うので、A列をNo. (狭く), B列を質問, C列を回答に再調整
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 80
# サマリー表は2列なので、C列は空でOK。代わりにサマリー表を再描画
for r in range(4, 14):
    ws.cell(row=r, column=3).fill = PatternFill(fill_type=None)
# 行高を調整
for r in range(18, 22):
    ws.row_dimensions[r].height = 90 if r > 18 else 22

# ============================================================
# Sheet 2: 開発目的とゴール
# ============================================================
ws = wb.create_sheet("2.目的とゴール")
write_title(ws, "開発目的とゴール", "なぜ作るのか・何を達成したいのか")

rows = [
    ["区分", "項目", "内容"],
    [
        "短期目的",
        "MVP完成と初期ユーザー検証",
        "AIで投資情報を自動キュレーションし、Discordに配信する仕組みを2週間で構築。"
        "10名のテストユーザー (個人投資家初級〜中級) に実利用してもらい、"
        "『情報の有用性・配信頻度・通知UX』に関する一次検証データを取得する。",
    ],
    [
        "中期目的",
        "補助金申請用エビデンスの確立",
        "MVPで得たユーザー利用実績・継続率・反応データを材料に、"
        "事業性証明 (実需と継続性) を補助金申請書類に反映できる形で整備する。"
        "MVP段階での『使われる証跡』が、申請通過と次フェーズの開発資金確保のレバーになる。",
    ],
    [
        "長期目的",
        "投資判断支援SaaSとしての事業化",
        "β版で配信先をLINE/X等に拡張、有料化を試行。"
        "本番では全銘柄AI分析・自社AIモデル化まで拡張し、月額課金SaaSとして事業化を目指す。"
        "投資助言ではなく『意思決定のための情報提供』としてのポジショニングを維持。",
    ],
    [
        "ゴール定義 (MVP)",
        "成功基準",
        "①ニュース収集→AI処理→Discord配信が無人で1日2〜3回安定稼働する。"
        "②10名のテストユーザーがDiscordに参加し、配信を受け取れる。"
        "③ジャンル別の重要度ルールに基づき、緊急ニュース (例: 中東情勢→石油系) は即時配信できる。"
        "④運用コストが月5万円以内に収まる。",
    ],
    [
        "非ゴール (MVPでやらないこと)",
        "スコープ外",
        "・投資助言・売買推奨機能 (法的リスクのため)"
        "・全銘柄を網羅する分析"
        "・自社AIモデルの学習"
        "・有料課金システム"
        "・モバイルアプリ専用UI (Discord/Web確認のみ)",
    ],
]
apply_table(ws, 4, rows[0], rows[1:], col_widths=[18, 28, 80])
for r in range(5, 10):
    ws.row_dimensions[r].height = 95

# ============================================================
# Sheet 3: 開発スコープと方向性
# ============================================================
ws = wb.create_sheet("3.スコープと方向性")
write_title(ws, "開発スコープと方向性", "MVPで何をどこまで作るか")

rows = [
    ["カテゴリ", "対象", "MVPでの扱い", "理由・補足"],
    ["データ収集", "経済ニュースサイト (3〜5媒体)", "実装", "Yahoo!ファイナンス・東洋経済・ロイター・Bloomberg日本版・日経 (RSS優先)。APIで取れない媒体はスクレイピング。"],
    ["データ収集", "SNS (X, Reddit等)", "対象外", "MVPでは情報源を限定し、品質を優先。β版以降で検討。"],
    ["AI処理", "要約 (タイトル＋3行要約)", "実装", "OpenAI API (gpt-4o-mini or gpt-4o)。プロンプト固定化＋JSON出力で品質を安定化。"],
    ["AI処理", "ジャンル分類", "実装", "石油・金融・テック・地政学・エネルギー等の固定タグ体系。"],
    ["AI処理", "重要度スコアリング (1〜5)", "実装", "ジャンル別の閾値で配信頻度を制御。緊急(=5)は即時配信。"],
    ["AI処理", "リスク抽出・リード/ラグ分析", "簡易版のみ", "高度な相関分析はβ版以降。MVPは『関連ジャンルの頻度を一時的に上げる』ロジックのみ。"],
    ["配信", "Discord Bot 配信", "実装", "ジャンル別チャンネル＋全体チャンネル。1日2〜3回の定時 + 緊急即時。"],
    ["配信", "LINE / X / Web Push", "対象外 (抽象化のみ)", "Notifierインターフェースを抽象化し、本番で差し替え可能な構造にしておく。"],
    ["運用", "管理画面", "最小限 (CLI＋環境変数)", "ジャンル別配信頻度はYAML/DBで管理。Web UIは作らない。"],
    ["運用", "ユーザー管理", "対象外", "Discordサーバー参加=テストユーザー登録、で代用。"],
    ["品質", "プロンプト固定・JSON出力", "実装", "AI出力のブレ・ハルシネーション・ノイズ混入に対する基本対策。"],
    ["品質", "失敗時リトライ・ログ", "実装", "外部API障害でもパイプラインを止めない。Sentry/CloudWatch等は本番で導入。"],
    ["将来拡張", "配信先抽象化", "実装", "Notifier基底クラスを切る。LINE/X/Webhookを後で足せる構造。"],
    ["将来拡張", "自社AIモデル", "対象外", "β版以降。OpenAI APIで十分な品質が出るかをまずMVPで検証。"],
]
apply_table(ws, 4, rows[0], rows[1:], col_widths=[14, 30, 22, 60])

# ============================================================
# Sheet 4: 現状分析と開発開始方針
# ============================================================
ws = wb.create_sheet("4.現状分析と着手方針")
write_title(ws, "現状分析と開発開始方針", "現在のMVP状況を踏まえ、どこから着手するか")

rows = [
    ["区分", "項目", "現状認識 / 方針"],
    ["前提", "現在のフェーズ", "ヒアリングベースでは『MVP着手前 (要件整理段階)』。会話内では既にMVPがある場合も想定されるが、本提案は 0→1 構築前提で記載。既存コードがある場合は読み解きフェーズを追加。"],
    ["着手前", "要件・データソースの確定", "対象媒体 (3〜5サイト) を確定し、利用規約・robots.txt・APIライセンスを精査。スクレイピング可否を切り分け、コスト/法的リスクをマッピング。"],
    ["着手前", "ジャンル体系・配信ルールの初期設計", "ジャンル (石油・金融・地政学・テック等) と、ジャンルごとの配信頻度・重要度閾値を初期設計。クライアント承認を得てから実装に入る。"],
    ["Week1 着手内容", "データ収集パイプライン", "RSS/API取得→正規化→重複排除→DB保存のパイプラインを最優先で実装。スクレイピング対象は薄く広く確保し、無料tierの範囲で安定動作を優先。"],
    ["Week1 着手内容", "AI処理 (要約・分類・スコア)", "OpenAI API (gpt-4o-mini) でJSON強制出力。プロンプトはバージョン管理し、後で差し替え可能に。"],
    ["Week2 着手内容", "Discord Bot 配信", "ジャンル別チャンネル + 全体チャンネル。定時バッチ + 緊急即時のディスパッチャを実装。"],
    ["Week2 着手内容", "テストユーザー受入", "Discordサーバーを立ち上げ、10名招待。配信フォーマットの初期フィードバックを収集。"],
    ["Week3 (任意)", "改善・調整", "配信頻度・要約品質・ノイズ対策の調整。本番に向けたコスト最適化 (API呼出回数の削減、キャッシュ導入)。"],
    ["並行作業", "観点別の改善優先度設定", "①要約精度 / ②配信タイミング / ③ユーザーの理解しやすさ の3軸で課題を分類し、影響度の高いものから改善。"],
    ["並行作業", "Discordチャンネル設計", "・要件整理 / ・バグ報告 / ・改善案 をチャンネル分離し、コミュニケーションコストを下げる。クライアント・テストユーザー双方が参加。"],
    ["コミュニケーション", "進行スタイル", "テキストベース非同期。週1〜2回の同期ミーティング (任意)。1日2〜4時間稼働で継続的に関与可能。"],
]
apply_table(ws, 4, rows[0], rows[1:], col_widths=[16, 28, 80])
for r in range(5, 16):
    ws.row_dimensions[r].height = 70

# ============================================================
# Sheet 5: 運用コスト見積
# ============================================================
ws = wb.create_sheet("5.運用コスト見積")
write_title(ws, "運用コスト見積 (月額)", "サーバー・API等のランニングコスト ※開発費は含まず")

# MVP段階 (テストユーザー10名)
ws.cell(row=4, column=1, value="■ MVP段階 (テストユーザー10名・1日2〜3回配信)").font = Font(
    name="Yu Gothic", size=12, bold=True, color="1F4E78"
)
mvp_rows = [
    ["項目", "推奨構成", "月額(下限)", "月額(上限)", "備考"],
    ["サーバー", "AWS EC2 t3.small もしくは VPS (さくら/ConoHa 2GB)", 3000, 8000, "常時稼働。バッチのみならLambda化で更に安価可"],
    ["DB", "PostgreSQL (RDS db.t3.micro / VPS同居)", 0, 5000, "VPS同居なら追加費用ゼロ。RDSなら最小構成"],
    ["ストレージ", "S3もしくはVPSローカル (10〜50GB)", 0, 2000, "ニュース原文＋AI出力を保管。圧縮で大幅削減可"],
    ["OpenAI API", "gpt-4o-mini 中心 / 1日500〜2000記事処理", 5000, 20000, "要約＋分類＋スコア。プロンプト圧縮とキャッシュで削減可"],
    ["ニュースAPI", "NewsAPI 無料tier + 公式RSS + スクレイピング", 0, 0, "MVPは無料の範囲で構成。商用APIは本番で検討"],
    ["Discord", "Bot (無料)", 0, 0, "Discord Botは無料"],
    ["監視・ログ", "Sentry無料tier / CloudWatch最小", 0, 2000, "本格運用は本番で"],
    ["ドメイン・SSL", "（任意）", 0, 1000, "MVPでは不要、Web UI追加時のみ"],
    ["合計目安", "", 8000, 38000, "実運用は月15,000〜25,000円に収まる想定"],
]
apply_table(ws, 6, mvp_rows[0], mvp_rows[1:], col_widths=[20, 38, 12, 12, 60])
# 数値セルを通貨書式に
for r in range(7, 16):
    for c in (3, 4):
        ws.cell(row=r, column=c).number_format = "¥#,##0"
        ws.cell(row=r, column=c).alignment = RIGHT

# β版段階
ws.cell(row=18, column=1, value="■ β版段階 (ユーザー100名・1日5〜10回配信・10〜30媒体)").font = Font(
    name="Yu Gothic", size=12, bold=True, color="1F4E78"
)
beta_rows = [
    ["項目", "推奨構成", "月額(下限)", "月額(上限)", "備考"],
    ["サーバー", "AWS EC2 t3.medium + ALB / VPS 4GB×2", 10000, 25000, "配信並列化・冗長化"],
    ["DB", "RDS db.t3.small + バックアップ", 8000, 18000, "履歴保持と検索性能の向上"],
    ["ストレージ", "S3 100〜500GB", 1500, 5000, ""],
    ["OpenAI API", "gpt-4o + gpt-4o-mini ハイブリッド / 1日5000〜20000記事", 25000, 80000, "重要記事のみgpt-4oで再要約"],
    ["ニュースAPI", "商用API契約 (NewsAPI Business / 個別契約)", 5000, 30000, "正式契約により法的安全性とSLAを確保"],
    ["配信API", "LINE Messaging API / X API", 0, 15000, "送信数によりLINEは従量課金 / X API有料化"],
    ["監視・ログ", "Sentry / Datadog / CloudWatch", 3000, 10000, ""],
    ["ドメイン・SSL", "独自ドメイン + 証明書", 1000, 2000, ""],
    ["合計目安", "", 53500, 185000, "実運用は月8万〜15万円が現実的レンジ"],
]
apply_table(ws, 20, beta_rows[0], beta_rows[1:], col_widths=[20, 38, 12, 12, 60])
for r in range(21, 30):
    for c in (3, 4):
        ws.cell(row=r, column=c).number_format = "¥#,##0"
        ws.cell(row=r, column=c).alignment = RIGHT

# 本番段階
ws.cell(row=32, column=1, value="■ 本番段階 (ユーザー1,000名以上・全銘柄分析・複数チャネル)").font = Font(
    name="Yu Gothic", size=12, bold=True, color="1F4E78"
)
prod_rows = [
    ["項目", "推奨構成", "月額(下限)", "月額(上限)", "備考"],
    ["サーバー", "ECS/Fargate or EKS + ALB / RDS Multi-AZ", 40000, 120000, "オートスケール構成"],
    ["DB", "RDS db.t3.medium 〜 large + リードレプリカ", 25000, 60000, ""],
    ["ストレージ", "S3 1TB+ / Glacier アーカイブ", 5000, 20000, ""],
    ["OpenAI API / 自社モデル", "ハイブリッド or ファインチューニング自社モデル", 80000, 250000, "段階的に自社モデルへ移行しコスト削減"],
    ["ニュースAPI", "複数商用契約", 30000, 100000, ""],
    ["配信API", "LINE / X / Webhook / Email", 10000, 50000, ""],
    ["監視・ログ", "Datadog / Sentry / PagerDuty", 15000, 40000, ""],
    ["合計目安", "", 205000, 640000, "ユーザー数と記事処理量に強く比例"],
]
apply_table(ws, 34, prod_rows[0], prod_rows[1:], col_widths=[20, 38, 12, 12, 60])
for r in range(35, 43):
    for c in (3, 4):
        ws.cell(row=r, column=c).number_format = "¥#,##0"
        ws.cell(row=r, column=c).alignment = RIGHT

# コスト最適化メモ
ws.cell(row=45, column=1, value="■ コスト最適化のポイント").font = Font(
    name="Yu Gothic", size=12, bold=True, color="1F4E78"
)
notes = [
    "・OpenAI API: gpt-4o-mini を主力にし、重要度スコア=5の記事のみ gpt-4o で再要約することで API 費を 1/3〜1/5 に圧縮可能。",
    "・キャッシュ戦略: 同一URLは1度しか要約しない。タイトルハッシュで重複排除。",
    "・スクレイピング併用: 商用APIに頼り切らず、RSS+軽量スクレイピングで母数を確保。",
    "・サーバー: バッチ実行のみならLambda/Cloud Functions化で常時稼働コストをゼロに近づけられる。",
    "・配信: LINEは従量課金なのでスコア上位のみに絞る運用が効率的。",
]
for i, n in enumerate(notes, 47):
    ws.cell(row=i, column=1, value=n).font = NORMAL_FONT
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

# ============================================================
# Sheet 6: ビジネス優位性
# ============================================================
ws = wb.create_sheet("6.ビジネス優位性")
write_title(ws, "ビジネス優位性", "本プロジェクトの強みと差別化要因")

rows = [
    ["観点", "優位性", "根拠・補足"],
    ["市場性", "個人投資家層の情報過多問題に直接アプローチ", "投資系メディアは多いが、ジャンル別重要度フィルタを自動化したサービスは少ない。初級〜中級層の『何を見るべきか分からない』を解く。"],
    ["技術", "OpenAI APIによる短期立ち上げ", "自社モデル開発は不要。MVPは2週間で立ち上がり、検証→改善のループが早い。"],
    ["スケーラビリティ", "配信先を抽象化", "Discord→LINE→X→独自アプリ、と配信チャネルを差し替え可能な構造。市場の反応に応じてチャネル戦略を変えられる。"],
    ["事業化", "補助金活用と段階的事業化", "MVPで実績→補助金申請→β版開発→有料化、という段階設計。資金調達と開発を連動させられる。"],
    ["差別化", "投資助言ではなく『情報整理』に特化", "金融商品取引法の『助言・代理業』登録が不要なポジションを取れる。法務リスクが低くスピード重視で進められる。"],
    ["プロダクト", "ジャンル×重要度のマトリクスUX", "単なるニュース羅列ではなく『今読むべきものだけ』が届く設計。ユーザーの認知負荷を下げる。"],
    ["拡張性", "AIプロンプト・ロジックの差し替えが容易", "ロジックを外部化することで、市場・ユーザー反応に応じてキュレーション方針を柔軟に変えられる。"],
    ["コスト構造", "従量課金型の薄いインフラ", "MVP段階で月数万円、ユーザー増加に応じてリニアに増えない構造 (キャッシュ・閾値制御で抑制)。"],
    ["データ資産", "ニュース×ジャンル×反応の蓄積", "蓄積したデータは将来の自社AIモデル学習・有料分析レポートの種になる。長期的には資産価値が発生。"],
]
apply_table(ws, 4, rows[0], rows[1:], col_widths=[18, 38, 70])

# ============================================================
# Sheet 7: マイルストーン (段階別ゴール・作業内容)
# ============================================================
ws = wb.create_sheet("7.マイルストーン")
write_title(ws, "マイルストーン", "段階別ゴールと作業内容")

rows = [
    ["フェーズ", "期間", "ゴール", "主な作業内容", "成果物"],
    [
        "Phase 0: 要件整理",
        "0.5週",
        "対象媒体・ジャンル・配信ルールの初期合意",
        "・媒体リスト確定 (3〜5サイト)\n・ジャンル分類設計 (石油/金融/地政学/テック等)\n・ジャンル別配信頻度・重要度ルール初期案\n・利用規約・API契約条件の精査",
        "要件整理ドキュメント / ジャンル設計書",
    ],
    [
        "Phase 1: データ収集基盤",
        "Week1 前半",
        "ニュースが自動で集まりDBに保存される",
        "・RSS/API取得モジュール\n・スクレイピングモジュール (必要媒体のみ)\n・正規化・重複排除\n・PostgreSQLスキーマ設計\n・cronスケジューリング",
        "データ収集パイプライン (1時間ごとに自動実行)",
    ],
    [
        "Phase 2: AI処理",
        "Week1 後半",
        "収集したニュースが要約・分類・スコアリングされる",
        "・OpenAI APIラッパー (リトライ・JSON強制)\n・要約プロンプト v1\n・ジャンル分類プロンプト v1\n・重要度スコア (1〜5) ロジック\n・関連ジャンル頻度の一時引き上げロジック",
        "AI処理モジュール / プロンプトv1",
    ],
    [
        "Phase 3: Discord配信",
        "Week2 前半",
        "ジャンル別・全体・緊急の3系統で自動配信される",
        "・Discord Botセットアップ\n・ジャンル別チャンネル設計\n・定時配信スケジューラ\n・緊急即時配信トリガ\n・配信フォーマット (見出し/要約/リンク)",
        "Discord Bot / 配信フォーマット v1",
    ],
    [
        "Phase 4: テストユーザー受入",
        "Week2 後半",
        "10名のテストユーザーが実際に配信を受け取る",
        "・Discordサーバー設計 (チャンネル分離)\n・招待・オンボーディング案内\n・フィードバック収集チャンネル運営\n・初期挙動の監視",
        "稼働中のDiscordサーバー / 運用ガイド",
    ],
    [
        "Phase 5: 改善 (任意)",
        "Week3",
        "初期フィードバックを反映し、運用品質を上げる",
        "・要約品質の改善 (プロンプト調整)\n・配信頻度の調整\n・コスト最適化 (キャッシュ/閾値)\n・障害復旧手順の整備",
        "改善版v1.1 / 運用手順書",
    ],
    [
        "Phase 6: β版 (将来)",
        "1〜2ヶ月",
        "ユーザー100名・配信先拡張・補助金申請",
        "・配信先拡張 (LINE/X)\n・管理画面の追加\n・ユーザーごとの設定\n・有料tierの試験提供\n・補助金申請書類用エビデンス整理",
        "β版 / 申請用レポート",
    ],
    [
        "Phase 7: 本番 (将来)",
        "3ヶ月以降",
        "全銘柄AI分析・自社モデル化・SaaS化",
        "・自社AIモデル学習 (ファインチューニング)\n・全銘柄ニュース処理\n・課金システム\n・SLA運用",
        "本番SaaS",
    ],
]
apply_table(ws, 4, rows[0], rows[1:], col_widths=[20, 12, 35, 60, 32])
for r in range(5, 13):
    ws.row_dimensions[r].height = 130

# ============================================================
# Sheet 8: 機能実装計画
# ============================================================
ws = wb.create_sheet("8.機能実装計画")
write_title(ws, "機能実装計画", "段階別に実装する機能の一覧")

rows = [
    ["機能カテゴリ", "機能名", "MVP", "β版", "本番", "備考"],
    ["データ収集", "RSS取得", "○", "○", "○", "Yahoo!ファイナンス・ロイター等"],
    ["データ収集", "公式API取得", "○", "○", "○", "NewsAPI / 各社公式"],
    ["データ収集", "スクレイピング", "△", "○", "○", "MVPは限定。利用規約遵守"],
    ["データ収集", "SNS取得 (X/Reddit)", "−", "△", "○", "情報源拡張"],
    ["データ収集", "重複排除", "○", "○", "○", "URL/タイトルハッシュ"],
    ["データ収集", "正規化", "○", "○", "○", "DateTime/エンコーディング統一"],
    ["AI処理", "タイトル＋3行要約", "○", "○", "○", "gpt-4o-mini"],
    ["AI処理", "ジャンル分類", "○", "○", "○", "固定タグ体系"],
    ["AI処理", "重要度スコア (1〜5)", "○", "○", "○", "ジャンル別閾値"],
    ["AI処理", "リスク抽出", "△", "○", "○", "MVPは簡易"],
    ["AI処理", "リード/ラグ分析", "−", "△", "○", "関連銘柄推定"],
    ["AI処理", "自社モデル (FT)", "−", "−", "○", "本番で順次置換"],
    ["AI処理", "プロンプトバージョン管理", "○", "○", "○", "差し替え可能に"],
    ["配信", "Discord Bot配信", "○", "○", "○", "ジャンル別/全体/緊急"],
    ["配信", "定時配信スケジューラ", "○", "○", "○", "ジャンルごとに頻度"],
    ["配信", "緊急即時配信", "○", "○", "○", "重要度=5でトリガ"],
    ["配信", "LINE Messaging連携", "−", "○", "○", "Notifier抽象化済"],
    ["配信", "X (Twitter) 投稿", "−", "○", "○", "Notifier抽象化済"],
    ["配信", "Webhook連携", "−", "○", "○", "外部連携用"],
    ["配信", "Email配信", "−", "△", "○", "ダイジェスト用"],
    ["管理", "ジャンル設定 (YAML/DB)", "○", "○", "○", "MVPはYAML可"],
    ["管理", "Web管理画面", "−", "○", "○", "β版以降"],
    ["管理", "ユーザー管理", "−", "○", "○", "β版以降"],
    ["管理", "課金システム", "−", "△", "○", "本番で本格化"],
    ["品質・運用", "リトライ・エラーハンドリング", "○", "○", "○", ""],
    ["品質・運用", "ログ・監視", "△", "○", "○", "MVPは最小、本番でDatadog等"],
    ["品質・運用", "アラート", "−", "○", "○", ""],
    ["品質・運用", "オートスケール", "−", "△", "○", ""],
    ["品質・運用", "バックアップ", "△", "○", "○", "MVPは手動でも可"],
    ["分析", "配信反応の集計", "−", "○", "○", "リアクション/開封率"],
    ["分析", "ユーザー行動分析", "−", "△", "○", ""],
    ["分析", "AI出力評価ダッシュボード", "−", "△", "○", ""],
]
apply_table(ws, 4, rows[0], rows[1:], col_widths=[18, 32, 8, 8, 8, 40])
# ○△−のセル中央寄せ
for r in range(5, 5 + len(rows) - 1):
    for c in (3, 4, 5):
        ws.cell(row=r, column=c).alignment = CENTER

# 凡例
ws.cell(row=5 + len(rows) + 1, column=1, value="凡例: ○=実装する / △=部分的に実装 / −=対象外").font = Font(
    name="Yu Gothic", size=9, italic=True
)

# ============================================================
# Sheet 9: UI設計
# ============================================================
ws = wb.create_sheet("9.UI設計")
write_title(ws, "UI設計", "段階別の画面・UI構成")

# MVPのUI
ws.cell(row=4, column=1, value="■ MVP段階のUI (Discord中心 / 管理UIは最小)").font = Font(
    name="Yu Gothic", size=12, bold=True, color="1F4E78"
)
mvp_ui = [
    ["UI種別", "対象ユーザー", "画面/構成", "機能", "備考"],
    ["Discord (テストユーザー向け)", "10名のテストユーザー", "ジャンル別チャンネル", "ジャンルごとの定時配信を受信", "石油/金融/地政学/テック/全体 等"],
    ["Discord (テストユーザー向け)", "10名のテストユーザー", "全体チャンネル (#all)", "全ジャンル横断の重要ニュース", ""],
    ["Discord (テストユーザー向け)", "10名のテストユーザー", "緊急チャンネル (#urgent)", "重要度=5の即時配信", ""],
    ["Discord (フィードバック)", "全員", "#要件整理 / #バグ報告 / #改善案", "コミュニケーション用", "情報の混在を防ぐ"],
    ["管理", "開発者・クライアント", "YAML/環境変数", "ジャンル別配信頻度・閾値の設定", "Web UIは作らない"],
    ["管理", "開発者", "CLI (Python script)", "手動配信・再処理・状態確認", ""],
]
apply_table(ws, 6, mvp_ui[0], mvp_ui[1:], col_widths=[28, 22, 28, 38, 30])

# β版のUI
ws.cell(row=15, column=1, value="■ β版のUI (Web管理画面 + 複数チャネル配信)").font = Font(
    name="Yu Gothic", size=12, bold=True, color="1F4E78"
)
beta_ui = [
    ["UI種別", "対象ユーザー", "画面/構成", "機能", "備考"],
    ["Web管理画面", "管理者", "ダッシュボード", "配信状況・処理件数・コスト概算", ""],
    ["Web管理画面", "管理者", "ジャンル管理", "ジャンル追加/編集/閾値設定", ""],
    ["Web管理画面", "管理者", "プロンプト管理", "プロンプトのバージョン管理・差し替え", ""],
    ["Web管理画面", "管理者", "ユーザー管理", "ユーザー一覧・配信先設定", ""],
    ["ユーザー向け Web (任意)", "登録ユーザー", "ニュース閲覧画面", "配信されたニュースを後から閲覧", "Discord履歴の補助"],
    ["ユーザー向け Web (任意)", "登録ユーザー", "個人設定", "ジャンル選択・配信先 (Discord/LINE/X)", ""],
    ["配信", "登録ユーザー", "Discord", "MVPと同様", ""],
    ["配信", "登録ユーザー", "LINE", "Messaging APIによるプッシュ", ""],
    ["配信", "登録ユーザー", "X", "公式アカウントから投稿", ""],
]
apply_table(ws, 17, beta_ui[0], beta_ui[1:], col_widths=[28, 22, 28, 38, 30])

# 本番のUI
ws.cell(row=29, column=1, value="■ 本番のUI (SaaS化・有料機能)").font = Font(
    name="Yu Gothic", size=12, bold=True, color="1F4E78"
)
prod_ui = [
    ["UI種別", "対象ユーザー", "画面/構成", "機能", "備考"],
    ["LP / 申込", "見込み客", "ランディングページ", "サービス紹介・料金プラン・登録導線", ""],
    ["ユーザー向け Web", "有料会員", "ダッシュボード", "重要ニュース一覧・銘柄ウォッチリスト・履歴", ""],
    ["ユーザー向け Web", "有料会員", "個人設定", "ジャンル/配信先/通知時間/有料プラン", ""],
    ["ユーザー向け Web", "有料会員", "分析レポート", "週次/月次のジャンル別動向レポート", "有料tier"],
    ["管理画面", "運営", "課金管理", "Stripe等との連携", ""],
    ["管理画面", "運営", "AI出力評価", "プロンプト品質モニタリング", ""],
    ["管理画面", "運営", "障害監視", "Datadog等の組み込み", ""],
    ["配信", "有料会員", "Discord/LINE/X/Email", "全チャネル対応", ""],
]
apply_table(ws, 31, prod_ui[0], prod_ui[1:], col_widths=[28, 22, 28, 38, 30])

# UI設計方針
ws.cell(row=42, column=1, value="■ UI設計方針").font = Font(
    name="Yu Gothic", size=12, bold=True, color="1F4E78"
)
policy = [
    "・MVP段階ではユーザー向けUIは作らず、Discordをそのまま画面とみなす。これにより検証速度を最大化。",
    "・配信フォーマットは『見出し / 重要度バッジ / 3行要約 / ジャンルタグ / 元記事リンク』の固定構成で読み手の認知負荷を最小化。",
    "・管理は YAML+CLI で済ませ、Web 管理画面は β 版で初めて作る。",
    "・本番では LP / ダッシュボード / 個人設定 / レポート / 課金 を順次拡張。",
    "・配信先抽象化は MVP から仕込んでおき、本番までUIではなくバックエンドの差し替えで対応する。",
]
for i, n in enumerate(policy, 44):
    ws.cell(row=i, column=1, value=n).font = NORMAL_FONT
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

# ============================================================
# 保存
# ============================================================
wb.save(OUT_PATH)
print(f"OK: {OUT_PATH}")
